"""
Trillium V2 — Excel Generator
Genera file Excel (.xlsx) con la parts list e i pesi stimati.
Output worksheets: Summary, PartsList, Log.
"""

import io
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# STILI EXCEL
# ============================================================

if HAS_OPENPYXL:
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")
    SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")

    DATA_FONT = Font(name="Calibri", size=10)
    DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
    NUMBER_ALIGN = Alignment(horizontal="right", vertical="center")

    WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    GROUP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    GROUP_FONT = Font(name="Calibri", size=10, bold=True)

    TOTAL_FONT = Font(name="Calibri", size=12, bold=True)
    TOTAL_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    TOTAL_FONT_WHITE = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )


def generate_excel(estimation_result, filename: Optional[str] = None) -> io.BytesIO:
    """
    Genera file Excel con i risultati della stima.

    Args:
        estimation_result: EstimationResult dall'estimator
        filename: Nome file (opzionale, per naming convention)

    Returns:
        BytesIO contenente il file .xlsx
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl è richiesto per la generazione Excel. "
                          "Installa con: pip install openpyxl")

    wb = openpyxl.Workbook()

    # 1. Worksheet Summary
    _create_summary_sheet(wb, estimation_result)

    # 2. Worksheet PartsList
    _create_parts_list_sheet(wb, estimation_result)

    # 3. Worksheet Log
    _create_log_sheet(wb, estimation_result)

    # Rimuovi il foglio vuoto di default se è ancora lì
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Salva in BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_filename(estimation_result) -> str:
    """Genera il nome file seguendo la naming convention."""
    family = estimation_result.params.get("pump_family", "UNKNOWN")
    ts = estimation_result.timestamp.strftime("%Y%m%d-%H%M")
    job_id = estimation_result.job_id
    return f"PartsList_{family}_{ts}_{job_id}.xlsx"


# ============================================================
# WORKSHEET: SUMMARY
# ============================================================

def _create_summary_sheet(wb, result):
    ws = wb.create_sheet("Summary", 0)

    # Titolo
    ws.merge_cells("A1:F1")
    ws["A1"] = "TRILLIUM V2 — Weight Estimation Report"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Job ID: {result.job_id} — {result.timestamp.strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    # Parametri Input
    row = 4
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "PARAMETRI INPUT"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    row += 1

    params_labels = [
        ("Famiglia Pompa", result.params.get("pump_family", "")),
        ("Nq (Velocità Specifica)", result.params.get("nq", "")),
        ("Scale Factor (f)", result.params.get("scale_factor", "")),
        ("Pressione Progetto (bar)", result.params.get("pressure", "")),
        ("Temperatura Progetto (°C)", result.params.get("temperature", "")),
        ("Materiale", result.params.get("material", "")),
        ("Flange Rating", result.params.get("flange_rating", "")),
        ("Spessore Parete (mm)", result.params.get("wall_thickness", "")),
        ("Numero Stadi", result.params.get("num_stages", 1)),
        ("Aspirazione (pollici)", result.params.get("suction_size_inch", "")),
        ("Mandata (pollici)", result.params.get("discharge_size_inch", "")),
    ]

    for label, value in params_labels:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = DATA_FONT
        ws[f"B{row}"].border = THIN_BORDER
        row += 1

    row += 1

    # Riepilogo Risultati
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "RIEPILOGO RISULTATI"
    ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    row += 1

    summary_data = [
        ("Peso Totale Stimato (kg)", f"{result.total_weight_kg:.1f}"),
        ("Componenti Stimati", f"{len([c for c in result.components if c.is_estimated])}/{len(result.components)}"),
        ("Confidenza Alta", len([c for c in result.components if c.confidence == "alta"])),
        ("Confidenza Media", len([c for c in result.components if c.confidence == "media"])),
        ("Confidenza Bassa", len([c for c in result.components if c.confidence == "bassa"])),
        ("Warning Totali", len(result.warnings)),
    ]

    for label, value in summary_data:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = DATA_FONT
        ws[f"B{row}"].border = THIN_BORDER
        row += 1

    # Pompa di riferimento
    if result.ref_pump_info:
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = "POMPA DI RIFERIMENTO"
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="2F5496")
        row += 1
        for key, value in result.ref_pump_info.items():
            if key != "components":
                ws[f"A{row}"] = str(key)
                ws[f"A{row}"].font = Font(name="Calibri", size=10, bold=True)
                ws[f"B{row}"] = str(value)
                ws[f"B{row}"].font = DATA_FONT
                row += 1

    # Warning
    if result.warnings:
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = "⚠ WARNING"
        ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True, color="C00000")
        row += 1
        for w in result.warnings:
            ws[f"A{row}"] = f"• {w}"
            ws[f"A{row}"].font = DATA_FONT
            ws[f"A{row}"].fill = WARN_FILL
            row += 1

    # Larghezza colonne
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35


# ============================================================
# WORKSHEET: PARTS LIST
# ============================================================

def _create_parts_list_sheet(wb, result):
    ws = wb.create_sheet("PartsList")

    # Header
    headers = [
        "Componente", "Gruppo", "Metodo Calcolo",
        "Peso Rif. (kg)", "Materiale Rif.", "Sorgente",
        "Peso Stimato (kg)", "Fattori Applicati",
        "Confidenza", "Note", "Warning",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Dati
    current_group = None
    row = 2

    for comp in result.components:
        # Riga separatore per nuovo gruppo
        if comp.group != current_group:
            current_group = comp.group
            ws.merge_cells(f"A{row}:{get_column_letter(len(headers))}{row}")
            ws[f"A{row}"] = f"  {current_group}"
            ws[f"A{row}"].font = GROUP_FONT
            ws[f"A{row}"].fill = GROUP_FILL
            row += 1

        data = comp.to_dict()
        for col, header in enumerate(headers, 1):
            value = data.get(header, "")
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            if header in ("Peso Rif. (kg)", "Peso Stimato (kg)"):
                cell.alignment = NUMBER_ALIGN
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.0'

            elif header == "Confidenza":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if value == "alta":
                    cell.fill = OK_FILL
                elif value == "media":
                    cell.fill = WARN_FILL
                elif value == "bassa":
                    cell.fill = ERROR_FILL
            else:
                cell.alignment = DATA_ALIGN

        row += 1

    # Riga totale
    row += 1
    ws[f"A{row}"] = "TOTALE"
    ws[f"A{row}"].font = TOTAL_FONT_WHITE
    ws[f"A{row}"].fill = TOTAL_FILL
    for col in range(2, len(headers) + 1):
        ws.cell(row=row, column=col).fill = TOTAL_FILL

    ws[f"G{row}"] = result.total_weight_kg
    ws[f"G{row}"].font = TOTAL_FONT_WHITE
    ws[f"G{row}"].fill = TOTAL_FILL
    ws[f"G{row}"].number_format = '#,##0.0'
    ws[f"G{row}"].alignment = NUMBER_ALIGN

    # Larghezza colonne
    col_widths = [35, 15, 22, 14, 16, 25, 16, 30, 12, 30, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze panes
    ws.freeze_panes = "A2"


# ============================================================
# WORKSHEET: LOG
# ============================================================

def _create_log_sheet(wb, result):
    ws = wb.create_sheet("Log")

    # Header
    ws["A1"] = "Timestamp"
    ws["B1"] = "Evento"
    for col in range(1, 3):
        ws.cell(row=1, column=col).font = HEADER_FONT
        ws.cell(row=1, column=col).fill = HEADER_FILL
        ws.cell(row=1, column=col).alignment = HEADER_ALIGN

    # Dati log
    for i, entry in enumerate(result.log_entries, 2):
        # Separa timestamp dal messaggio se presente
        if entry.startswith("[") and "]" in entry:
            ts = entry[1:entry.index("]")]
            msg = entry[entry.index("]") + 2:]
        else:
            ts = ""
            msg = entry

        ws[f"A{i}"] = ts
        ws[f"A{i}"].font = Font(name="Consolas", size=9, color="808080")
        ws[f"B{i}"] = msg
        ws[f"B{i}"].font = Font(name="Consolas", size=9)

        # Colora in base al tipo
        if "WARNING" in msg or "⚠" in msg:
            ws[f"B{i}"].fill = WARN_FILL
        elif "ERRORE" in msg or "✗" in msg:
            ws[f"B{i}"].fill = ERROR_FILL
        elif "✓" in msg:
            ws[f"B{i}"].fill = OK_FILL

    # Larghezza colonne
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 100

    # Freeze panes
    ws.freeze_panes = "A2"
